import boto3
import logging
import math
import os
import openpyxl
import re
from datetime import datetime
from io import BytesIO
from collections import defaultdict
from typing import Dict, List, Any, Optional

logger = logging.getLogger()
logger.setLevel(logging.INFO)


class ExcelProcessor:
    dynamodb = boto3.resource("dynamodb")
    TABLE_NAME = os.environ.get("TABLE_NAME", "AdaProjectTable")
    table = dynamodb.Table(TABLE_NAME)


    @staticmethod
    def _safe_str(v: Any) -> str:
        return "" if v is None else str(v).strip()

    @staticmethod
    def _safe_int(v: Any, default: int = 0) -> int:
        try:
            if v is None or str(v).strip() == "":
                return default
            return int(str(v).strip())
        except Exception:
            return default


    @staticmethod
    def get_workbook_from_s3(bucket: str, key: str) -> openpyxl.Workbook:
        s3 = boto3.client("s3")
        try:
            data = s3.get_object(Bucket=bucket, Key=key)["Body"].read()
        except Exception as e:
            logger.error(f"Error al descargar {bucket}/{key} de S3: {e}")
            raise
        try:
            return openpyxl.load_workbook(BytesIO(data), data_only=True)
        except openpyxl.utils.exceptions.InvalidFileException as e:
            logger.error(f"Archivo XLSX inválido o corrupto: {e}")
            raise


    @staticmethod
    def format_date(value: Any) -> str:
        if value is None:
            return ""

        if hasattr(value, "strftime"):
            try:
                return value.strftime("%d/%m/%Y")
            except Exception:
                pass

        s = str(value).strip()
        if not s or s.lower() == "none":
            return ""

        s_iso = s.replace("T", " ").replace("Z", "")
        s_iso = re.sub(r"\.\d+$", "", s_iso)
        for pat in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s_iso, pat)
                return dt.strftime("%d/%m/%Y")
            except Exception:
                pass

        for pat in ("%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(s, pat)
                return dt.strftime("%d/%m/%Y")
            except Exception:
                pass

        for pat in ("%m/%d/%Y %H:%M:%S", "%m-%d-%Y %H:%M:%S", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                dt = datetime.strptime(s, pat)
                return dt.strftime("%d/%m/%Y")
            except Exception:
                pass

        return s


    @staticmethod
    def compute_semester(start_date_str: str, credit_date_str: str) -> str:
        try:
            start_date = datetime.strptime(start_date_str, "%d/%m/%Y")
            credit_date = datetime.strptime(credit_date_str, "%d/%m/%Y")
        except Exception as ex:
            logger.warning(f"Error en conversión de fecha en compute_semester: {ex}")
            return "0"

        diff_months = (credit_date.year - start_date.year) * 12 + (credit_date.month - start_date.month)
        semester = math.floor(diff_months / 6) + 1
        return str(max(semester, 0))


    @staticmethod
    def _validate_header(header: List[Any], expected_cols: List[str], section_name: str) -> Dict[str, int]:
        normalized = [ExcelProcessor._safe_str(h) for h in header]
        col_idx = {col.strip(): idx for idx, col in enumerate(normalized)}
        for col in expected_cols:
            if col not in col_idx:
                logger.warning(f"Columna esperada no encontrada en {section_name}: {col}")
        return col_idx


    @staticmethod
    def process_activities_sheet(wb: openpyxl.Workbook, sheet_name: str = "activities") -> Dict[str, List[Dict[str, Any]]]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Hoja '{sheet_name}' no encontrada en el libro.")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        header = [ExcelProcessor._safe_str(cell) for cell in rows[0]]
        expected = [
            "ID_COMIENZO", "NOMBRE_COMIENZO", "ID_DICTADO", "ID_MATERIA",
            "DESCRIPCION_MATERIA", "id_codigo", "Fecha Obtención Credito",
            "TIPO_CREDITO_CTA_CTE_ACD", "RESULTADO_CTA_CTE_ACD",
            "CALIFICACION_CTA_CTE_ACD", "TIPO_DE_OBTENCION"
        ]
        idx = ExcelProcessor._validate_header(header, expected, "activities")

        activities_by_student: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

        for row in rows[1:]:
            if row is None:
                continue
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            student_id = ExcelProcessor._safe_str(row[idx.get("id_codigo")])
            if not student_id:
                continue

            credit_raw = row[idx.get("Fecha Obtención Credito")]
            credit_date = ExcelProcessor.format_date(credit_raw)

            activity = {
                "ID_MATERIA": ExcelProcessor._safe_str(row[idx.get("ID_MATERIA")]),
                "DESCRIPCION_MATERIA": ExcelProcessor._safe_str(row[idx.get("DESCRIPCION_MATERIA")]),
                "Fecha Obtención Credito": credit_date,
                "TIPO_CREDITO_CTA_CTE_ACD": ExcelProcessor._safe_str(row[idx.get("TIPO_CREDITO_CTA_CTE_ACD")]),
                "RESULTADO_CTA_CTE_ACD": ExcelProcessor._safe_str(row[idx.get("RESULTADO_CTA_CTE_ACD")]),
                "CALIFICACION_CTA_CTE_ACD": ExcelProcessor._safe_str(row[idx.get("CALIFICACION_CTA_CTE_ACD")]),
                "TIPO_DE_OBTENCION": ExcelProcessor._safe_str(row[idx.get("TIPO_DE_OBTENCION")]),
            }
            activities_by_student[student_id].append(activity)

        return activities_by_student

    @staticmethod
    def process_global_info_sheet(wb: openpyxl.Workbook, sheet_name: str = "global_info") -> Dict[str, Dict[str, Any]]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Hoja '{sheet_name}' no encontrada en el libro.")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        header = [ExcelProcessor._safe_str(cell) for cell in rows[0]]
        expected = [
            "id_codigo",
            "NOMBRE_TITULO",
            "PLAN_TITULO",
            "Comienzo",
            "Graduación"
        ]
        idx = ExcelProcessor._validate_header(header, expected, "global info")

        student_info: Dict[str, Dict[str, Any]] = {}

        for row in rows[1:]:
            if row is None:
                continue
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            student_id = ExcelProcessor._safe_str(row[idx.get("id_codigo")])
            if not student_id:
                continue

            start_date = ExcelProcessor.format_date(row[idx.get("Comienzo")])
            graduation_date = ExcelProcessor.format_date(row[idx.get("Graduación")])

            info = {
                "title": ExcelProcessor._safe_str(row[idx.get("NOMBRE_TITULO")]),
                "plan": ExcelProcessor._safe_str(row[idx.get("PLAN_TITULO")]),
                "start_date": start_date,
                "graduation_date": graduation_date
            }
            student_info[student_id] = info

        return student_info


    @classmethod
    def process_xlsx(cls, bucket: str, key: str) -> None:
        wb = cls.get_workbook_from_s3(bucket, key)
        activities = cls.process_activities_sheet(wb, sheet_name="activities")
        globals_ = cls.process_global_info_sheet(wb, sheet_name="global_info")

        for student_id, acts in activities.items():
            record = {
                "PK": f"STUDENT#{student_id}",
                "SK": f"ACTIVITIES#{student_id}",
                "activities": acts,
                "info": globals_.get(student_id, {})
            }
            cls.table.put_item(Item=record)
            logger.info(f"Guardado en DynamoDB: {student_id}")
